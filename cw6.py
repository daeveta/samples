# f = open('cw4.py', 'a+')
# f.write('# this is out test from other script \n')
# f.seek(0)
# #f.close()
# #f = open('05-02.py', 'r')
# print(f.read())
#
# with open('cw4.py', 'a') as f:
#     f.write('print("some text")\n')
#
# with open('cw4.py', 'a') as f:
#     print(f.read())
# import json
# from os.path import exists
#
# file_name = "cw4.classworkscriptdataexample"
# if exists(file_name):
#     print('file exists')
#     students_list = []
#     with open(file_name) as file:
#         students_list = json.loads(file.read())
# else:
#     students_list = [
#         'daeveta', 'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
#             'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
#             'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
#     ]
#     with open(file_name, 'w') as file:
#         file.write(json.dumps(students_list))
#
# print(students_list)

# import csv
# import random
#
#
# students_list = [
#          'daeveta', 'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
#              'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
#              'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
# ]
# csv_data = []
# for i, student in enumerate(students_list):
#     csv_data.append((i+1, student, random.choice(students_list)))
#
# # print(csv_data)
#
# with open('students.csv', 'w') as file:
#     wr = csv.writer(file)
#     for student_info in csv_data:
#         wr.writerow(student_info)
#
# with open('students.csv', 'r') as file:
#     rr = csv.reader(file)
#     for student_info in rr:
#         if student_info:
#             print(student_info)
import csv
import random
import openpyxl

students_list = [
         'daeveta', 'dontlivr', 'DSTroj', 'eevgenygmir', 'GrigoriyBaranchuk',
             'Karinastriker', 'maminadochka', 'Muskulan', 'neverloaded96',
             'psenkevich', 'quackee', 'TonyMahouney', 'VitaliaKulenko'
]

csv_data = []
for i, student in enumerate(students_list):
    csv_data.append((i+1, student, random.choice(students_list)))

wb = openpyxl.Workbook()
sheet = wb.active
# c1 = sheet.cell(row=1, column=1)
# c1.value = 'ankit'

for students_data in csv_data:
    for column, cell in enumerate(students_data):
        data_cell = sheet.cell(row=students_data[0], column=column+1)
        data_cell.value = cell
# установить pip
#